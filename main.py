from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
import time

usuarios = ["", "", "", ""]
passwords = ["", "", "", ""]
excels = ["sgfal_cantabria.xlsx", "sgfal_extremadura.xlsx", "sgfal_murcia.xlsx", "sgfal_la_rioja.xlsx"]

for i in range(4):
    service = Service(executable_path="chromedriver.exe")
    driver = webdriver.Chrome(service = service)
    driver.get("https://serviciostelematicosext.hacienda.gob.es/SGCIEF/Inventario_CCAA/secciones/Login.aspx")

    usuario = usuarios[i]
    password = passwords[i]
    excel = excels[i]
        
    input_usuario = driver.find_element(By.ID, "txt_usuario")
    input_password = driver.find_element(By.ID, "txt_pass")
    
    input_usuario.send_keys(usuario)
    input_password.send_keys(password + Keys.ENTER)
    
    boton_lista_entes = driver.find_element(By.ID, "lnkTodosEntes")
    boton_lista_entes.send_keys(Keys.ENTER)
    
    
    ## ESTOY EN LA TABLA DE CADA ENTE
    # Encuentra la tabla por su ID
    table = driver.find_element(By.ID, 'TablaDetalle')
    
    # Encuentra todas las filas dentro del cuerpo de la tabla (tbody)
    rows = table.find_elements(By.XPATH, './/tr')
    
    # Itera a través de las filas y obtiene los valores de la columna "Código de Ente"
    codigo_ente_values = []
    
    for row in rows[1:]:  # Saltar la primera fila que contiene los encabezados
        # Encuentra la tercera columna de cada fila
        try:
            codigo_ente = row.find_elements(By.XPATH, './td')[2].text
            codigo_ente_values.append(codigo_ente)
            print(codigo_ente)
        except IndexError:
            # Si la fila no tiene suficientes columnas, ignora esa fila
            continue
    
    
    # Crear un nuevo libro de Excel
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Eliminar la hoja predeterminada
    
    # Iterar a través de los valores de "Código de Ente"
    for codigo in codigo_ente_values:
        try:
            print("Iteración para: " + codigo)
            # Encuentra el enlace por el texto específico
            link_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f"//a[text()='{codigo}']"))
            )
    
            # Guardar el link en una variable
            link = link_element.get_attribute('href')
    
            # Hacer clic en el enlace
            ActionChains(driver).move_to_element(link_element).click().perform()
    
            #Le doy a "Listado de todos los datos del ente"
            boton_lista_entes = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "lnkListarEnte"))
            )
            boton_lista_entes.send_keys(Keys.ENTER)
    
            # Esperar a que la página cargue
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, 'body'))  # Espera a que el cuerpo de la nueva página se cargue
            )
    
            # Obtener el HTML de la página actual
            page_html = driver.page_source
    
            # Crear una nueva hoja en el libro de Excel con el nombre del código
            sheet = workbook.create_sheet(title=codigo)
    
            # Usar BeautifulSoup para procesar el HTML
            soup = BeautifulSoup(page_html, 'html.parser')
            lines = soup.prettify().split('\n')
    
            # Escribir cada línea del HTML en una fila de la hoja de Excel
            for i, line in enumerate(lines, start=1):
                sheet.cell(row=i, column=1, value=line)
    
            print(page_html[:200])
    
            # Volver a la página principal
            driver.back()
            driver.back()
    
            # Esperar a que la página principal cargue de nuevo
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, 'TablaDetalle'))
            )
        except Exception as e:
            print(f"Error al procesar el ente '{codigo}': {str(e)}")
    
    # Guardar el libro de Excel
    workbook.save(excel)
    print("FIN "+ excel)
    driver.quit()





#for item in downloaded_htmls:
#    print(f"Código: {item['codigo']}")
#    print(f"URL: {item['url']}")
#    print(f"HTML: {item['html'][:400]}...")  # Imprime solo los primeros 400 caracteres

print('fin total')
#time.sleep(10)
#driver.quit()
